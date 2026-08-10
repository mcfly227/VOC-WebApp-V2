"""Generates the Aggregate Emissions Report workbook in the v10 layout:

  1. Aggregate Emissions Report  (monthly + 12-mo rolling, tons; DBE/EB monthly in lbs)
  2. Aggregate Material Use      (monthly gallons)
  3. Isobutyl Acetate Report     (daily, lbs/8-hr shift vs 153.6 limit)
  4-6. Daily Use per coating line
  7. Material Content Report     (lbs/gal per product)
  8. As Applied VOC

Historical rows (through the frozen horizon) are written verbatim from the
imported v10 data; newer months are computed from usage logs. All values are
formatted to 4 decimal places per the EGLE work instruction.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session
from . import models, calc

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
FMT4 = "0.0000"
HDR = Font(bold=True)
FILL = PatternFill("solid", fgColor="DDE6F0")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Border(bottom=Side(style="thin", color="AAAAAA"))


def _setting(db, key, default=None):
    s = db.get(models.Setting, key)
    return s.value if s else default


def generate_workbook(db: Session, year: int, month: int) -> bytes:
    hap_columns = _setting(db, "hap_columns", [])
    mu_hap_columns = _setting(db, "mu_hap_columns", [])
    series = calc.monthly_series(db, year, month, hap_columns, mu_hap_columns)

    wb = Workbook()
    _tab1(wb.active, series, hap_columns)
    _tab2(wb.create_sheet(), series, mu_hap_columns)
    _tab3(wb.create_sheet(), db, series)
    for i, eu in enumerate(models.EMISSION_UNITS):
        _daily_tab(wb.create_sheet(), db, series, eu, 4 + i)
    _tab7(wb.create_sheet(), db)
    _tab8(wb.create_sheet(), db)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _head(ws, cells):
    for ref, text in cells.items():
        ws[ref] = text
        ws[ref].font = HDR
        ws[ref].alignment = CENTER


def _tab1(ws, series, hap_columns):
    ws.title = "1.Aggregate Emisions Report"
    _head(ws, {"A1": "PTI 183-15", "D1": "VOC Emissions",
               "M1": "CAS Specific Emissions", "T1": "HAPs Emissions"})
    ws["M2"] = "627930, 1119400, 106650"; ws["O2"] = "100414"; ws["Q2"] = "98828"
    ws["T2"] = "Aggregate HAPs"
    ws["M3"] = "Dibasic Ester"; ws["O3"] = "Ethylbenzene"; ws["Q3"] = "Cumene"
    ws["A4"] = "Equipment:"
    ws["D4"] = "EU-CoatingLine-01"; ws["F4"] = "EU-CoatingLine-02"
    ws["H4"] = "EU-CoatingLine-03"; ws["J4"] = "FG-Coating"
    ws["M4"] = "FG-Coating"; ws["O4"] = "FG-Coating"; ws["Q4"] = "FG-Facility"; ws["T4"] = "FG-Facility"
    units = ["Units", "Limit", "", "Tons", "40 TPY", "Tons", "40 TPY", "Tons", "10 TPY",
             "Tons", "89.9 TPY", "", "Lbs", "2.9 TPY", "Lbs", "2.9 TPY", "Tons", "1.4 TPY",
             "", "Tons", "8.9 TPY"]
    for j, v in enumerate(units, start=1):
        ws.cell(row=5, column=j, value=v)
    labels = ["Year", "Month", "", "VOC Monthly", "VOC 12-Mo Rolling",
              "VOC Monthly", "VOC 12-Mo Rolling", "VOC Monthly", "VOC 12-Mo Rolling",
              "VOC Monthly", "VOC 12-Mo Rolling", "",
              "Dibasic Ester Monthly", "Dibasic Ester 12-Mo Rolling",
              "Ethylbenzene Monthly", "Ethylbenzene 12-Mo Rolling",
              "Cumene Monthly", "Cumene 12-Mo Rolling", "",
              "Aggregate HAPs Monthly", "Aggregate HAPS 12-Mo Rolling"]
    for j, v in enumerate(labels, start=1):
        c = ws.cell(row=6, column=j, value=v)
        c.font = HDR; c.alignment = CENTER; c.fill = FILL
    col = 22
    for h in hap_columns:
        ws.cell(row=2, column=col, value=h["cas"]).font = HDR
        ws.cell(row=3, column=col, value=h["name"]).font = HDR
        ws.cell(row=5, column=col, value="Tons")
        ws.cell(row=5, column=col + 1, value="8.9 TPY")
        ws.cell(row=6, column=col, value="Monthly").fill = FILL
        ws.cell(row=6, column=col + 1, value="12-Mo Rolling").fill = FILL
        col += 2

    row = 7
    prev_year = None
    for rec in series:
        if rec["year"] != prev_year:
            ws.cell(row=row, column=1, value=rec["year"]).font = HDR
            prev_year = rec["year"]
        ws.cell(row=row, column=2, value=MONTH_NAMES[rec["month"] - 1])
        em, rl = rec["emissions"], rec["rolling"]
        vals = [(4, em.get("voc_eu1", 0)), (5, rl["voc_eu1"]),
                (6, em.get("voc_eu2", 0)), (7, rl["voc_eu2"]),
                (8, em.get("voc_eu3", 0)), (9, rl["voc_eu3"]),
                (10, em.get("voc_eu1", 0) + em.get("voc_eu2", 0) + em.get("voc_eu3", 0)),
                (11, rl["voc_eu1"] + rl["voc_eu2"] + rl["voc_eu3"]),
                (13, em.get("dbe_lbs", 0)), (14, rl["dbe_tons"]),
                (15, em.get("eb_lbs", 0)), (16, rl["eb_tons"]),
                (17, em.get("cumene_tons", 0)), (18, rl["cumene_tons"]),
                (20, em.get("agghap_tons", 0)), (21, rl["agghap_tons"])]
        for cidx, v in vals:
            c = ws.cell(row=row, column=cidx, value=round(v or 0.0, 10))
            c.number_format = FMT4
        col = 22
        hm = em.get("hap_tons") or {}
        hr = rl.get("hap_tons") or {}
        for idx in range(len(hap_columns)):
            c1 = ws.cell(row=row, column=col, value=round(hm.get(str(idx), 0.0), 10))
            c2 = ws.cell(row=row, column=col + 1, value=round(hr.get(str(idx), 0.0), 10))
            c1.number_format = FMT4; c2.number_format = FMT4
            col += 2
        row += 1
    ws.freeze_panes = "C7"
    for letter in ("A", "B"):
        ws.column_dimensions[letter].width = 8
    for j in range(3, 22):
        ws.column_dimensions[ws.cell(row=6, column=j).column_letter].width = 14


def _tab2(ws, series, mu_hap_columns):
    ws.title = "2.Aggregate Material Use"
    _head(ws, {"A1": "PTI 183-15", "D1": "VOC Containing Materials",
               "I1": "CAS Specific Containing Materials", "L1": "HAPs Containing Materials"})
    ws["I2"] = "627930, 1119400, 106650"; ws["J2"] = "100414"; ws["K2"] = "98828"
    ws["L2"] = "Aggregate HAPs"
    ws["I3"] = "Dibasic Ester"; ws["J3"] = "Ethylbenzene"; ws["K3"] = "Cumene"
    ws["D3"] = "EU-CoatingLine-01"; ws["E3"] = "EU-CoatingLine-02"
    ws["F3"] = "EU-CoatingLine-03"; ws["G3"] = "FG-Facility"
    for j, v in enumerate(["Units", "", "", "Gals", "Gals", "Gals", "Gals", "",
                           "Gals", "Gals", "Gals", "Gals"], start=1):
        ws.cell(row=5, column=j, value=v)
    for j, v in enumerate(["Year", "Month", "", "Monthly", "Monthly", "Monthly",
                           "Monthly", "", "Monthly", "Monthly", "Monthly", "Monthly"], start=1):
        c = ws.cell(row=6, column=j, value=v); c.font = HDR; c.fill = FILL
    col = 13
    for h in mu_hap_columns:
        ws.cell(row=2, column=col, value=h["cas"]).font = HDR
        ws.cell(row=3, column=col, value=h["name"]).font = HDR
        ws.cell(row=5, column=col, value="Gals")
        ws.cell(row=6, column=col, value="Monthly").fill = FILL
        col += 1
    row = 7
    prev_year = None
    for rec in series:
        if rec["year"] != prev_year:
            ws.cell(row=row, column=1, value=rec["year"]).font = HDR
            prev_year = rec["year"]
        ws.cell(row=row, column=2, value=MONTH_NAMES[rec["month"] - 1])
        mu = rec["material_use"]
        fg = mu.get("gals_eu1", 0) + mu.get("gals_eu2", 0) + mu.get("gals_eu3", 0)
        vals = [(4, mu.get("gals_eu1", 0)), (5, mu.get("gals_eu2", 0)),
                (6, mu.get("gals_eu3", 0)), (7, fg),
                (9, mu.get("gals_dbe", 0)), (10, mu.get("gals_eb", 0)),
                (11, mu.get("gals_cumene", 0)), (12, mu.get("gals_agghap", 0))]
        for cidx, v in vals:
            c = ws.cell(row=row, column=cidx, value=round(v or 0.0, 10))
            c.number_format = FMT4
        hg = mu.get("hap_gals") or {}
        for idx in range(len(mu_hap_columns)):
            c = ws.cell(row=row, column=13 + idx, value=round(hg.get(str(idx), 0.0), 10))
            c.number_format = FMT4
        row += 1
    ws.freeze_panes = "C7"


def _horizon(db):
    s = db.get(models.Setting, "frozen_through")
    y, m = s.value if s else (2026, 6)
    return f"{y}{m:02d}32"   # any YYYYMMDD in/before that month sorts below this


def _tab3(ws, db, series):
    ws.title = "3.IsobutylAcetate Report(Daily)"
    _head(ws, {"A1": "PTI 183-15", "C1": "Isobutyl Acetate Containing Materials",
               "C2": "FG-Coating"})
    for j, v in enumerate(["Year", "Date", "Product Number", "Material Use",
                           "Isobutyl Acetate Content", "Isobutyl Acetate Emissions"], start=1):
        c = ws.cell(row=3, column=j, value=v); c.font = HDR; c.fill = FILL
    for j, v in enumerate(["YYYY", "YYYYMMDD", "", "(Gals)", "(Lbs/gal)",
                           "(Lbs/8-hour Shift)"], start=1):
        ws.cell(row=4, column=j, value=v)
    ws.cell(row=5, column=6, value="Limit: 153.6 lbs/8-hr")
    row = 6
    prev_year = None
    hz = _horizon(db)
    for fr in db.query(models.FrozenDailyRow).filter_by(sheet="IBA").order_by(models.FrozenDailyRow.id):
        r = fr.row
        if r[1] > hz:
            continue
        yr = r[1][:4]
        ws.cell(row=row, column=1, value=int(yr) if yr != str(prev_year) else None)
        if yr != str(prev_year):
            prev_year = int(yr)
        ws.cell(row=row, column=2, value=r[1])
        ws.cell(row=row, column=3, value=r[2])
        for j, v in [(4, r[3]), (5, r[4]), (6, r[5])]:
            try:
                c = ws.cell(row=row, column=j, value=float(v)); c.number_format = FMT4
            except (TypeError, ValueError):
                ws.cell(row=row, column=j, value=v)
        row += 1
    for rec in series:
        if rec["frozen"] or not rec["iba"]:
            continue
        for r in rec["iba"]:
            yr = int(r["date"][:4])
            if yr != prev_year:
                ws.cell(row=row, column=1, value=yr); prev_year = yr
            ws.cell(row=row, column=2, value=r["date"])
            ws.cell(row=row, column=3, value=r["product"])
            for j, v in [(4, r["gallons"]), (5, r["content"]), (6, r["lbs_per_8hr"])]:
                c = ws.cell(row=row, column=j, value=round(v, 10)); c.number_format = FMT4
            row += 1


def _daily_tab(ws, db, series, eu, n):
    ws.title = f"{n}.Daily Use {eu}"
    _head(ws, {"A1": "PTI 183-15", "B1": "Material Use Report", "B2": eu})
    for j, v in enumerate(["Year", "Date", "Coating Type", "Product Number",
                           "Part Type", "Material Use"], start=1):
        c = ws.cell(row=3, column=j, value=v); c.font = HDR; c.fill = FILL
    for j, v in enumerate(["YYYY", "YYYYMMDD", "Basecoat, Clearcoat, Primer, Hardener, or Solvent",
                           "", "Automotive or Non-Automotive Specialty", "Gals"], start=1):
        ws.cell(row=4, column=j, value=v)
    row = 5
    prev_year = None
    hz = _horizon(db)
    for fr in db.query(models.FrozenDailyRow).filter_by(sheet=eu).order_by(models.FrozenDailyRow.id):
        r = fr.row
        if r[1] > hz:
            continue
        yr = r[1][:4]
        if yr != str(prev_year):
            ws.cell(row=row, column=1, value=int(yr)); prev_year = int(yr)
        ws.cell(row=row, column=2, value=r[1])
        ws.cell(row=row, column=3, value=r[2])
        ws.cell(row=row, column=4, value=r[3])
        ws.cell(row=row, column=5, value=r[4])
        try:
            c = ws.cell(row=row, column=6, value=float(r[5])); c.number_format = FMT4
        except (TypeError, ValueError):
            ws.cell(row=row, column=6, value=r[5])
        row += 1
    for rec in series:
        if rec["frozen"] or not rec["daily"]:
            continue
        for r in rec["daily"].get(eu, []):
            yr = int(r["date"][:4])
            if yr != prev_year:
                ws.cell(row=row, column=1, value=yr); prev_year = yr
            ws.cell(row=row, column=2, value=r["date"])
            ws.cell(row=row, column=3, value=r["coating_type"])
            ws.cell(row=row, column=4, value=r["product"])
            ws.cell(row=row, column=5, value=r["part_type"])
            c = ws.cell(row=row, column=6, value=round(r["gallons"], 10))
            c.number_format = FMT4
            row += 1
    for letter, w in [("B", 12), ("C", 12), ("D", 22), ("E", 24), ("F", 12)]:
        ws.column_dimensions[letter].width = w


def _tab7(ws, db):
    ws.title = "7.Material Content Report"
    _head(ws, {"A1": "PTI 183-15", "B1": "Material Content Report", "B2": "FG-Coating"})
    cas_hdr = ["", "", "", "98828", "627930", "1119400", "106650", "100414"]
    for j, v in enumerate(cas_hdr, start=1):
        ws.cell(row=3, column=j, value=v).font = HDR
    names = ["", "HAP Content", "VOC Content", "Cumene Content", "Dimethyl Adipate Content",
             "Dimethyl Glutarate Content", "Dimethyl Succinate Content", "Ethylbenzene Content"]
    for j, v in enumerate(names, start=1):
        c = ws.cell(row=4, column=j, value=v); c.font = HDR; c.fill = FILL
    for j, v in enumerate(["Product Number", "Lbs/gal", "Lbs/gal", "Lbs/gal", "Lbs/gal",
                           "Lbs/gal", "Lbs/gal", "Lbs/gal"], start=1):
        ws.cell(row=5, column=j, value=v)
    row = 6
    for p in db.query(models.Product).order_by(models.Product.number):
        cc = p.cas_content()
        ov = p.content_override or {}
        vals = [p.eds_hap_content or ov.get("hap", 0.0),
                p.voc_content or ov.get("voc", 0.0),
                cc.get(models.CUMENE_CAS, 0.0),
                cc.get("627930", 0.0), cc.get("1119400", 0.0), cc.get("106650", 0.0),
                cc.get(models.EB_CAS, 0.0)]
        ws.cell(row=row, column=1, value=p.number)
        for j, v in enumerate(vals, start=2):
            c = ws.cell(row=row, column=j, value=round(v, 10)); c.number_format = "0.######"
        row += 1
    ws.column_dimensions["A"].width = 22
    for letter in "BCDEFGH":
        ws.column_dimensions[letter].width = 15


def _tab8(ws, db):
    ws.title = "8.As Applied VOC"
    _head(ws, {"A1": "PTI 183-15", "B1": "As-Applied VOC Report", "B2": "FG-Coating"})
    for j, v in enumerate(["", "As-Applied VOC", "Coating Category"], start=1):
        c = ws.cell(row=4, column=j, value=v); c.font = HDR; c.fill = FILL
    for j, v in enumerate(["Product Number", "Lbs/gal", "Automotive or non-Automotive"], start=1):
        ws.cell(row=5, column=j, value=v)
    row = 6
    prods = db.query(models.Product).order_by(models.Product.number).all()
    by_number = {p.number: p for p in prods}
    for p in prods:
        v = models.as_applied_voc_effective(p, by_number)
        if v is None:
            continue
        ws.cell(row=row, column=1, value=p.number)
        c = ws.cell(row=row, column=2, value=round(v, 4))
        c.number_format = "0.00"
        ws.cell(row=row, column=3, value=p.as_applied_category)
        row += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 28
